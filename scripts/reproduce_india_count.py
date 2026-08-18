#!/usr/bin/env python3
"""Reproduce India's English-language count from Census 2011 table C-17.

Usage:
    python reproduce_india_count.py /path/to/DDW-C17-0000.XLSX output.csv

The script restricts rows to state code 00 (India), then sums English as:
mother tongue + first subsidiary language + second subsidiary language.
"""

from __future__ import annotations

import csv
import hashlib
import sys
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHA256 = "cd74f457dbd62017c919e763e8f4e956b4f8c46c0920500907d949828ccfe673"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip().upper()


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: reproduce_india_count.py INPUT_XLSX OUTPUT_CSV")

    source_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    source_hash = sha256(source_path)
    if source_hash != EXPECTED_SHA256:
        raise SystemExit(
            "Unexpected workbook SHA-256. "
            f"Expected {EXPECTED_SHA256}, received {source_hash}."
        )

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    mother_tongue = 0
    first_subsidiary = 0
    second_subsidiary = 0

    for row in worksheet.iter_rows(min_row=7, values_only=True):
        state_code = normalize(row[0]).zfill(2)
        if state_code != "00":
            continue
        if normalize(row[3]) == "ENGLISH":
            mother_tongue += int(row[4] or 0)
        if normalize(row[8]) == "ENGLISH":
            first_subsidiary += int(row[9] or 0)
        if normalize(row[13]) == "ENGLISH":
            second_subsidiary += int(row[14] or 0)

    values = [
        ("english_mother_tongue", mother_tongue),
        ("english_first_subsidiary_language", first_subsidiary),
        ("english_second_subsidiary_language", second_subsidiary),
        ("english_total_up_to_three_languages", mother_tongue + first_subsidiary + second_subsidiary),
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as destination:
        writer = csv.writer(destination)
        writer.writerow(["measure", "persons", "source_sha256"])
        for measure, persons in values:
            writer.writerow([measure, persons, source_hash])

    for measure, persons in values:
        print(f"{measure}: {persons:,}")


if __name__ == "__main__":
    main()
